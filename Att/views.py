import datetime
import os

import uuid
import xlrd
import requests
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.shortcuts import render
from django.http import HttpResponse
from django.http import FileResponse
from django.views.decorators.csrf import csrf_exempt
# from pip._internal.resolution.resolvelib.resolver import Result

from Att import models
from common import models
from common.models import Project, Att, ProjectItem, DDuser, Group, UserGroups, UserPosition, Position, Deptment, \
    DeptmentUser

# Create your views here.
from djangoProject import settings
from django.utils import timezone
from openpyxl import load_workbook


# def __init__(self, itemName):
#     self.ItemName = itemName
#
# def learn(self):
#     print('%s is learning' % self.itemName

# @csrf_exempt
def up_file_excel(file):
    # 根name取 file 的值
    # logger.log().info('uplaod:%s' % file)
    # 创建upload文件夹
    furl = ""
    if not os.path.exists(settings.UPLOAD_ROOT):
        os.makedirs(settings.UPLOAD_ROOT)
    try:
        if file is None:
            return HttpResponse('请选择要上传的文件')
        # 循环二进制写入
        furl = settings.UPLOAD_ROOT + "/" + file.name
        print(furl)
        with open(settings.UPLOAD_ROOT + "/" + file.name, 'wb') as f:
            for i in file.readlines():
                f.write(i)
    except Exception as e:
        return HttpResponse(e)

    # return HttpResponse('上传成功')
    return furl


# 将excel数据写入mysql
# @staticmethod  #静态方法
# @classmethod#类方法
def wrdb(file, projectId):
    d1 = timezone.now()
    # print(myform)
    print("file:" + file)

    fdate = ProjectItem.objects.filter(projectId=projectId)
    if fdate:
        # 删除projectId明细
        df = ProjectItem.objects.filter(projectId=projectId)
        df.delete()
    wb = load_workbook(filename=settings.UPLOAD_ROOT + file)  # 打开文件,默认可读写，若有需要可以指定write_only和read_only为True
    ws = wb[wb.sheetnames[0]]  # 选择第一张sheet表
    rows = ws.max_row  # 获取表的最大行数
    columns = ws.max_column  # 获取表的最大列数

    column_heading = [ws.cell(row=1, column=x).value for x in range(1, columns + 1)]  # 读取excel第一行的值，写入list

    column_name = ['材料（设备）名称', '规格型号', '品牌', '单位', '数量']  # 数据库必需字段

    print("文件第一行title:" + str(column_heading))  # 文件第一行title

    if len([name for name in column_name if name not in column_heading]) == 0:  # 返回字段组成的list为空，则说明文件列标题包含MySQL需要的字段
        print(' - 检查完成，执行写入')
        # 判断Excel中各字段所在列号
        itemName = column_heading.index(column_name[0])  # 材料（设备）名称 - 位置
        Specs = column_heading.index(column_name[1])  # 规格型号 - 位置
        Brand = column_heading.index(column_name[2])  # 品牌 - 位置
        Unit = column_heading.index(column_name[3])  # 单位 - 位置
        Count = column_heading.index(column_name[4])  # 数量 - 位置

        if ws.cell(row=2, column=1).value == None:
            table_start_line = 3
        else:
            table_start_line = 2
        data = []
        itemList = []
        for row in range(table_start_line, rows + 1):
            for column in range(1, columns + 1):  # 因为从第1列开始，所以此处从1开始
                data.append(str(ws.cell(row=row, column=column).value))  # 以字符串形式保存数据到MySQL
            print("sss:" + str(data))
            print(data[itemName], data[Specs], data[Brand], data[Unit], data[Count])
            itemList.append(
                ProjectItem(projectId=projectId, itemName=data[itemName], Specs=data[Specs], Brand=data[Brand],
                            Unit=data[Unit], Count=data[Count], add_date=d1, up_date=d1))
            data = []
        print('itemList ', itemList)

        try:
            ProjectItem.objects.bulk_create(itemList)  # 使用bulk_create批量导入
            msg = '导入成功'
        except Exception as e:
            print('异常', e)
            msg = '导入失败'
    else:
        print('文件列标题不完全包含数据库需要的字段，请检查文件。')
        msg = '文件列标题不完全包含数据库需要的字段，请检查文件。'
    wb.close()  # 关闭excel

    return msg


def userdb(file):
    d1 = timezone.now()
    # print(myform)
    print("file:" + file)
    wb = load_workbook(filename=settings.UPLOAD_ROOT + file)  # 打开文件,默认可读写，若有需要可以指定write_only和read_only为True
    ws = wb[wb.sheetnames[0]]  # 选择第一张sheet表
    rows = ws.max_row  # 获取表的最大行数
    columns = ws.max_column  # 获取表的最大列数

    column_heading = [ws.cell(row=3, column=x).value for x in range(1, columns + 1)]  # 读取excel第三行的值，写入list

    column_name = ['员工ID', '姓名', '手机号', '部门', '职位']  # 数据库必需字段

    print("文件第三行title:" + str(column_heading))  # 文件第三行title

    if len([name for name in column_name if name not in column_heading]) == 0:  # 返回字段组成的list为空，则说明文件列标题包含MySQL需要的字段
        print(' - 检查完成，执行写入')
        # 判断Excel中各字段所在列号
        userid = column_heading.index(column_name[0])  # 员工ID
        name = column_heading.index(column_name[1])  # 姓名
        phone = column_heading.index(column_name[2])  # 手机号
        groupName = column_heading.index(column_name[3])  # 部门
        position = column_heading.index(column_name[4])  # 职位
        userPosition = '普通员工'

        if ws.cell(row=4, column=1).value == None:
            table_start_line = 5
        else:
            table_start_line = 4
        data = []
        itemList = []
        for row in range(table_start_line, rows + 1):
            for column in range(1, columns + 1):  # 因为从第1列开始，所以此处从1开始
                data.append(str(ws.cell(row=row, column=column).value))  # 以字符串形式保存数据到MySQL
            print("sss:" + str(data))
            print(data[userid], data[name], data[phone], data[groupName], data[position])
            if data[position] == '':
                userP = '普通员工'
            else:
                userP = data[position]
            ##插入用户信息到账号表
            uid = 0
            uname = ''
            ph = data[phone].split('-')
            try:
                password = make_password('nfc!@123$%')
                try:
                    ul = User.objects.get(username=ph[1])
                    uid = ul.id  # 获取新用户ID
                    dul = DDuser.objects.get(uid=ul.id)
                    uname = dul.name
                    print("用户账号存在跳过新增")
                except User.DoesNotExist:
                    uadd = User.objects.create(username=ph[1],
                                               password=password,
                                               is_superuser=1,
                                               last_name='',
                                               email='',
                                               is_staff=1,
                                               is_active=1,
                                               date_joined=d1,
                                               first_name='')
                    uid = uadd.id  # 获取新用户ID
                    uname = data[name]
                    print("新增用户账号表成功。用户名：" + str(uadd.username))
            except User.DoesNotExist:
                print("新增用户账号表失败。")
            ## 插入用户信息到账号表

            did = 0
            try:
                dn = Deptment.objects.get(deptName=data[groupName])
                print("部门存在,开始写入")
                try:
                    udc = DeptmentUser.objects.get(userID=uid, deptID=dn.deptID)
                    did = udc.deptID
                    print("该用户已存在部门中跳过新增" + str(udc.userName))
                except DeptmentUser.DoesNotExist:
                    dcord = DeptmentUser.objects.create(userID=uid, userName=uname, deptID=dn.deptID)
                    did = dcord.deptID
                    print("写入成功：" + str(dcord.userName))
            except Deptment.DoesNotExist:
                print("部门不存在，插入该部门。")
                dcord = Deptment.objects.create(deptName=data[groupName])
                did = dcord.deptID
                print("部门写入成功：" + str(dcord.deptName))
                try:
                    udc = DeptmentUser.objects.get(userID=uid, deptID=dcord.deptID)

                    if udc:
                        print("账号的用户组数据存在跳过新增" + str(udc.ID))
                        did = udc.deptID
                    else:
                        dncord = DeptmentUser.objects.create(userID=uid, deptID=udc.deptID)
                        did = dncord.deptID
                    print("该用户已存在部门中跳过新增" + str(udc.userName))
                except DeptmentUser.DoesNotExist:
                    dcord = DeptmentUser.objects.create(userID=uid, userName=uname, deptID=dcord.deptID)
                    did = dcord.deptID
                    print("写入成功：" + str(dcord.userName))

            gid = 0
            try:
                gn = Group.objects.get(groupName=data[groupName])
                print("用户组存在,开始写入")
                try:
                    ugc = UserGroups.objects.get(userID=uid, groupID=gn.groupID)
                    if ugc:
                        print("账号的用户组数据存在跳过新增" + str(ugc.ID))
                        gid = ugc.ID
                    else:
                        gncord = UserGroups.objects.create(userID=uid, groupID=gn.groupID)
                        gid = gncord.ID
                except UserGroups.DoesNotExist:
                    gncord = UserGroups.objects.create(userID=uid, groupID=gn.groupID)
                    gid = gncord.ID
                    print("写入成功。ID：" + str(gncord.ID))
            except Group.DoesNotExist:
                print("用户组不存在，写入无权限组别。")
                gnx = Group.objects.get(groupName='无权限组别')
                print("无权限组别ID：" + str(gnx.groupID))
                gncord = UserGroups.objects.create(userID=uid, groupID=gnx.groupID)
                gid = gncord.ID
                print("写入成功。ID：" + str(gncord.ID))

            try:
                pn = Position.objects.get(positionName=userP)
                try:
                    upc = UserPosition.objects.get(userID=uid, positionID=pn.positionID)
                    if upc:
                        print("账号的职务数据存在跳过新增")
                    else:
                        pncord = UserPosition.objects.create(userID=uid, positionID=pn.positionID, deptID=did)
                        print("写入成功。ID：" + str(pncord.positionID))
                except UserPosition.DoesNotExist:
                    pncord = UserPosition.objects.create(userID=uid, positionID=pn.positionID, deptID=did)
                    print("写入成功。ID：" + str(pncord.positionID))

            except Position.DoesNotExist:
                print("职务不存在，新增职务。")
                padd = Position.objects.create(positionName=userP)
                print("新增职务：" + userPosition + ",ID:" + str(padd.positionID))
                try:
                    upc = UserPosition.objects.get(userID=uid, positionID=padd.positionID)
                    print("账号的职务数据存在跳过新增")
                except UserPosition.DoesNotExist:
                    pncord = UserPosition.objects.create(userID=uid, positionID=padd.positionID)
                    print("写入成功。ID：" + str(pncord.positionID))
            try:
                du = DDuser.objects.get(name=data[name])
                print("用户明细表存在数据，跳过新增" + data[name])
            except DDuser.DoesNotExist:
                itemList.append(
                    DDuser(name=data[name], sys_level=0, is_sys=0, userid=data[userid], uid=uid, phone=ph[1]))
            data = []
        print('itemList ', itemList)

        try:
            DDuser.objects.bulk_create(itemList)  # 使用bulk_create批量导入
            msg = '导入成功'
        except Exception as e:
            print('异常', e)
            msg = '导入失败'
    else:
        print('文件列标题不完全包含数据库需要的字段，请检查文件。')
        msg = '文件列标题不完全包含数据库需要的字段，请检查文件。'
    wb.close()  # 关闭excel
    return msg


# @csrf_exempt
def upload(file, typeid, request):
    # 根name取 file 的值
    # file = request.FILES.get('file')
    # 取日期
    d1 = timezone.now()
    f_url = settings.UPLOAD_ROOT + "/" + str(d1.year) + "/" + str(d1.month)  # upload/年份/月份
    f_ext = file.name.split(".")[-1]
    out_url = ''
    # uuid.uuid1()　　基于MAC地址，时间戳，随机数来生成唯一的uuid，可以保证全球范围内的唯一性。
    #
    # 　　uuid.uuid2()　　算法与uuid1相同，不同的是把时间戳的前4位置换为POSIX的UID。不过需要注意的是python中没有基于DCE的算法，所以python的uuid模块中没有uuid2这个方法。
    #
    # 　　uuid.uuid3(namespace,
    #              name)　　通过计算一个命名空间和名字的md5散列值来给出一个uuid，所以可以保证命名空间中的不同名字具有不同的uuid，但是相同的名字就是相同的uuid了。【感谢评论区大佬指出】namespace并不是一个自己手动指定的字符串或其他量，而是在uuid模块中本身给出的一些值。比如uuid.NAMESPACE_DNS，uuid.NAMESPACE_OID，uuid.NAMESPACE_OID这些值。这些值本身也是UUID对象，根据一定的规则计算得出。
    #
    # 　　uuid.uuid4()　　通过伪随机数得到uuid，是有一定概率重复的
    #
    # 　　uuid.uuid5(namespace, name)　　和uuid3基本相同，只不过采用的散列算法是sha1

    # f_name = "xwl" + str(d1.year) + str(d1.month) + str(d1.day) + str(d1.second) + str(d1.microsecond)
    f_name = "xwl" + str(uuid.uuid1())
    print('f_url:%s' % f_url)
    print('uplaod:%s' % file)
    print('f_name:%s' % f_name)
    print('f_ext:%s' % f_ext)
    ip = request.META['REMOTE_ADDR']
    # import socket
    # # 函数 gethostname() 返回当前正在执行 Python 的系统主机名
    # res = socket.gethostbyname(socket.gethostname())
    # print(res)
    # ip = res
    print('ip:%s' % ip)
    # 创建upload文件夹
    if not os.path.exists(f_url):
        os.makedirs(f_url)
    try:
        if file is None:
            return HttpResponse('请选择要上传的文件')
        print(f_url + "/" + file.name)
        # 循环二进制写入
        with open(f_url + "/" + f_name + "." + f_ext, 'wb') as f:
            for i in file.readlines():
                f.write(i)
        f_size = os.path.getsize(f_url + "/" + f_name + "." + f_ext)
        out_url = "/" + str(d1.year) + "/" + str(d1.month) + "/" + f_name + "." + f_ext

        print('返回路径:' + out_url)
        # 附件参数写入 mysql
        now_time = datetime.datetime.now()
        print(str(now_time))
        record = Att.objects.create(TypeID=typeid,
                                    ParentID=0,
                                    FileUrl=out_url,
                                    Ext=f_ext,
                                    UserID=request.session['userid'],
                                    UserName=request.session['username'],
                                    Des=file.name,
                                    Tag="",
                                    Size=f_size,
                                    AddTime=now_time,
                                    Ip=ip)
        print('附件信息写入数据库:' + record.AttID)

        # 表格明细写入 mysql
        # wrdb(file.name)
    except Exception as e:
        # return HttpResponse(e)
        print('上传失败:' + out_url)

    # return HttpResponse('导入成功')
    # return HttpResponse(out_url)
    # print('返回路径:' + out_url)
    return out_url


def file_down(request):
    f = Att.objects.get(AttID=request.GET.get('attid'))
    file = open(settings.UPLOAD_ROOT + f.FileUrl, 'rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename=' + f.FileUrl
    print('附件信息:' + f.Des)
    return response
